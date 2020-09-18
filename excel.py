# Copyright (C) 2020 * Ltd. All rights reserved.
# author : Sanghyeon Jo <josanghyeokn@gmail.com>

import os
import openpyxl

class XLSX_Writer:
    def __init__(self, excel_path : str, class_names : list):
        if os.path.isfile(excel_path):
           raise ValueError 

        self.excel_path = excel_path
        
        self.book = openpyxl.Workbook()
        self.book.remove(self.book.get_sheet_by_name('Sheet'))
        self.sheet = self.book.create_sheet('Papers')

        self.start_ascii = 65
        self.row_index = 1

        self.max_length_of_column = {}

        self.update(class_names)

    def __call__(self, dataset : list):
        self.update(dataset)

    def update(self, dataset : list):
        for index, data in enumerate(dataset):
            column = chr(self.start_ascii + index)

            self.sheet['{}{}'.format(column, self.row_index)] = data

            try:
                length_of_data = len(data)
            except:
                length_of_data = len(str(data)) * 2

            try:
                self.max_length_of_column[column] = max(length_of_data, self.max_length_of_column[column])
            except KeyError:
                self.max_length_of_column[column] = length_of_data
            self.sheet.column_dimensions['{}'.format(column)].width = self.max_length_of_column[column]

        self.row_index += 1

    def close(self):
        self.book.save(self.excel_path)

class XLSX_Reader:
    def __init__(self, excel_path : str, class_names : list):
        if not os.path.isfile(excel_path):
           raise ValueError 

        self.book = openpyxl.load_workbook(excel_path)
        self.sheet = self.book.get_sheet_by_name('Papers')

        self.get_fn = lambda name: self.sheet[name].value

        self.dataset = []
        self.class_names = class_names

        self.start_ascii = 65
        self.row_index = 2

        self.update()

    def update(self):
        while True:

            data = []
            for index in range(len(self.class_names)):
                column = chr(self.start_ascii + index)
                data.append(self.get_fn('{}{}'.format(column, self.row_index)))

            if data[0] is None or len(data) == 0:
                break

            self.dataset.append(data)
            self.row_index += 1

if __name__ == '__main__':
    reader = XLSX_Reader('./test.xlsx', ['Name', 'Label'])

    for data in reader.dataset:
        print(data)